import tkinter as tk
from tkinter import ttk, messagebox
import re
import os
from openpyxl import Workbook, load_workbook

# -------------------- HÀM LƯU VÀ XỬ LÝ DỮ LIỆU --------------------

def save_to_excel(data, subjects):
    """Lưu thông tin vào file registration.xlsx.
       Mỗi môn học được chọn lưu thành 1 dòng riêng."""
    filename = "registration.xlsx"
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Thêm tiêu đề các cột (tuỳ chỉnh theo yêu cầu)
        ws.append(["Mã SV", "Họ tên", "Ngày sinh", "Email",
                   "Số ĐT", "Học kỳ", "Năm học", "Môn học"])
    
    for subject in subjects:
        row = [
            data["student_id"],
            data["ho_ten"],
            data["ngay_sinh"],
            data["email"],
            data["phone"],
            data["hoc_ky"],
            data["nam_hoc"],
            subject
        ]
        ws.append(row)
    wb.save(filename)

def register():
    """Hàm xử lý khi người dùng nhấn nút 'Đăng ký'."""
    student_id = entry_student_id.get().strip()
    ho_ten = entry_ho_ten.get().strip()
    ngay_sinh = entry_ngay_sinh.get().strip()
    email = entry_email.get().strip()
    phone = entry_phone.get().strip()
    hoc_ky = entry_hoc_ky.get().strip()
    nam_hoc = entry_nam_hoc.get().strip()
    
    # Lấy các môn học đã chọn
    selected_subjects = [sb for sb, var in subject_vars.items() if var.get() == 1]
    
    # Kiểm tra thông tin đã nhập
    if not (student_id and ho_ten and ngay_sinh and email and phone and hoc_ky and nam_hoc):
        messagebox.showerror("Lỗi", "Vui lòng nhập đầy đủ thông tin!")
        return
    
    # 1. Mã SV: 7 chữ số
    if not (student_id.isdigit() and len(student_id) == 7):
        messagebox.showerror("Lỗi", "Mã số sinh viên phải là 7 chữ số!")
        return
    
    # 2. Ngày sinh: định dạng dd/mm/yyyy
    if not re.match(r'^\d{2}/\d{2}/\d{4}$', ngay_sinh):
        messagebox.showerror("Lỗi", "Ngày sinh phải theo định dạng dd/mm/yyyy!")
        return
    
    # 3. Email: dùng regex để kiểm tra
    if not re.match(r'^[\w\.-]+@[\w\.-]+\.\w+$', email):
        messagebox.showerror("Lỗi", "Email không hợp lệ!")
        return
    
    # 4. Số điện thoại: 10 chữ số
    if not (phone.isdigit() and len(phone) == 10):
        messagebox.showerror("Lỗi", "Số điện thoại phải là 10 chữ số!")
        return
    
    # 5. Học kỳ: chỉ được nhập 1, 2 hoặc 3
    if hoc_ky not in ["1", "2", "3"]:
        messagebox.showerror("Lỗi", "Học kỳ chỉ được nhập 1, 2 hoặc 3!")
        return
    
    # 6. Năm học: chọn từ danh sách
    if nam_hoc not in ["2022-2023", "2023-2024", "2024-2025"]:
        messagebox.showerror("Lỗi", "Năm học không hợp lệ!")
        return
    
    # 7. Kiểm tra ít nhất 1 môn học
    if not selected_subjects:
        messagebox.showerror("Lỗi", "Vui lòng chọn ít nhất 1 môn học!")
        return
    
    # Nếu hợp lệ, lưu vào Excel
    data = {
        "student_id": student_id,
        "ho_ten": ho_ten,
        "ngay_sinh": ngay_sinh,
        "email": email,
        "phone": phone,
        "hoc_ky": hoc_ky,
        "nam_hoc": nam_hoc
    }
    save_to_excel(data, selected_subjects)
    messagebox.showinfo("Thành công", "Đăng ký thành công!")
    clear_form()

def clear_form():
    """Xoá nội dung trên form."""
    entry_student_id.delete(0, tk.END)
    entry_ho_ten.delete(0, tk.END)
    entry_ngay_sinh.delete(0, tk.END)
    entry_email.delete(0, tk.END)
    entry_phone.delete(0, tk.END)
    entry_hoc_ky.delete(0, tk.END)
    entry_nam_hoc.delete(0, tk.END)
    for var in subject_vars.values():
        var.set(0)

def exit_app():
    """Đóng ứng dụng."""
    root.destroy()

# -------------------- XÂY DỰNG GIAO DIỆN --------------------

root = tk.Tk()
root.title("Đăng ký học phần")
root.geometry("600x400")
root.config(bg="#90EE90")  # Màu xanh nhạt

# Tiêu đề lớn
label_title = tk.Label(root,
                       text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN",
                       font=("Arial", 16, "bold"),
                       fg="red",
                       bg="#90EE90")
label_title.pack(pady=5)

# Frame chứa các trường nhập liệu
frame_form = tk.Frame(root, bg="#90EE90")
frame_form.pack(pady=5, fill=tk.X)

# Mã số sinh viên
lbl_mssv = tk.Label(frame_form, text="Mã số sinh viên:", bg="#90EE90")
lbl_mssv.grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
entry_student_id = tk.Entry(frame_form, width=30)
entry_student_id.grid(row=0, column=1, padx=5, pady=2)

# Họ tên
lbl_ho_ten = tk.Label(frame_form, text="Họ tên:", bg="#90EE90")
lbl_ho_ten.grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
entry_ho_ten = tk.Entry(frame_form, width=30)
entry_ho_ten.grid(row=1, column=1, padx=5, pady=2)

# Ngày sinh
lbl_ngay_sinh = tk.Label(frame_form, text="Ngày sinh:", bg="#90EE90")
lbl_ngay_sinh.grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
entry_ngay_sinh = tk.Entry(frame_form, width=30)
entry_ngay_sinh.grid(row=2, column=1, padx=5, pady=2)

# Email
lbl_email = tk.Label(frame_form, text="Email:", bg="#90EE90")
lbl_email.grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
entry_email = tk.Entry(frame_form, width=30)
entry_email.grid(row=3, column=1, padx=5, pady=2)

# Số điện thoại
lbl_phone = tk.Label(frame_form, text="Số điện thoại:", bg="#90EE90")
lbl_phone.grid(row=4, column=0, sticky=tk.W, padx=5, pady=2)
entry_phone = tk.Entry(frame_form, width=30)
entry_phone.grid(row=4, column=1, padx=5, pady=2)

# Học kỳ
lbl_hoc_ky = tk.Label(frame_form, text="Học kỳ (1,2,3):", bg="#90EE90")
lbl_hoc_ky.grid(row=5, column=0, sticky=tk.W, padx=5, pady=2)
entry_hoc_ky = tk.Entry(frame_form, width=30)
entry_hoc_ky.grid(row=5, column=1, padx=5, pady=2)

# Năm học
lbl_nam_hoc = tk.Label(frame_form, text="Năm học:", bg="#90EE90")
lbl_nam_hoc.grid(row=6, column=0, sticky=tk.W, padx=5, pady=2)
entry_nam_hoc = tk.Entry(frame_form, width=30)
entry_nam_hoc.grid(row=6, column=1, padx=5, pady=2)
entry_nam_hoc.insert(0, "2022-2023")  # Gợi ý mặc định

# Môn học
lbl_mon_hoc = tk.Label(frame_form, text="Chọn môn học:", bg="#90EE90")
lbl_mon_hoc.grid(row=7, column=0, sticky=tk.W, padx=5, pady=2)

# Các checkbox môn học
subject_vars = {}
subjects = ["Lập trình Python", "Lập trình Java",
            "Công nghệ phần mềm", "Phát triển ứng dụng web"]

frame_subjects = tk.Frame(frame_form, bg="#90EE90")
frame_subjects.grid(row=7, column=1, padx=5, pady=2, sticky=tk.W)

# Để giống ảnh minh hoạ: sắp xếp 2 cột, 2 dòng
for i, subject in enumerate(subjects):
    var = tk.IntVar()
    chk = tk.Checkbutton(frame_subjects, text=subject, variable=var, bg="#90EE90")
    chk.grid(row=i//2, column=i%2, padx=5, pady=2, sticky=tk.W)
    subject_vars[subject] = var

# Frame chứa nút
frame_buttons = tk.Frame(root, bg="#90EE90")
frame_buttons.pack(pady=10)

btn_register = tk.Button(frame_buttons, text="Đăng ký", width=10, command=register)
btn_register.grid(row=0, column=0, padx=10)

btn_exit = tk.Button(frame_buttons, text="Thoát", width=10, command=exit_app)
btn_exit.grid(row=0, column=1, padx=10)

root.mainloop()
